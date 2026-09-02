from pathlib import Path

from openpyxl import load_workbook


def scan_excel_files(input_folder):
    """返回 input 目录中的所有 .xlsx 文件。"""
    return sorted(Path(input_folder).glob("*.xlsx"))


def get_sheet_count(file_path):
    """读取 Excel 文件并返回工作表数量。"""
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return len(workbook.sheetnames)
    finally:
        workbook.close()


def main(input_folder=None):
    """扫描并输出 Excel 文件信息。"""
    if input_folder is None:
        input_folder = Path(__file__).resolve().parent / "input"

    excel_files = scan_excel_files(input_folder)

    print("Excel 文件扫描工具")
    print("------------------")
    print(f"共发现 {len(excel_files)} 个 Excel 文件")

    if not excel_files:
        print("提示：input 文件夹中没有 .xlsx 文件。")

    for file in excel_files:
        size_kb = file.stat().st_size / 1024
        print()
        print(f"文件：{file.name}")
        print(f"大小：{size_kb:.2f} KB")

        try:
            sheet_count = get_sheet_count(file)
            print(f"Sheet数量：{sheet_count}")
        except Exception:
            print("读取失败：该文件可能已损坏，或不是有效的 Excel xlsx 文件。")


if __name__ == "__main__":
    main()
