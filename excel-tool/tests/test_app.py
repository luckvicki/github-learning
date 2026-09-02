import io
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from pathlib import Path

from openpyxl import Workbook


EXCEL_TOOL_DIR = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(EXCEL_TOOL_DIR))

import app


class ExcelToolTests(unittest.TestCase):
    def test_scan_excel_files_finds_only_xlsx_files(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_folder = Path(temp_dir)
            (input_folder / "first.xlsx").touch()
            (input_folder / "second.xlsx").touch()
            (input_folder / "ignored.txt").touch()

            files = app.scan_excel_files(input_folder)

            self.assertEqual(
                [file.name for file in files],
                ["first.xlsx", "second.xlsx"],
            )

    def test_get_sheet_count_returns_three(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            file_path = Path(temp_dir) / "three-sheets.xlsx"
            workbook = Workbook()
            workbook.active.title = "Sheet1"
            workbook.create_sheet("Sheet2")
            workbook.create_sheet("Sheet3")
            workbook.save(file_path)
            workbook.close()

            self.assertEqual(app.get_sheet_count(file_path), 3)

    def test_invalid_xlsx_does_not_stop_processing(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_folder = Path(temp_dir)
            (input_folder / "broken.xlsx").write_text(
                "not a valid Excel file",
                encoding="utf-8",
            )

            valid_file = input_folder / "valid.xlsx"
            workbook = Workbook()
            workbook.save(valid_file)
            workbook.close()

            output = io.StringIO()
            with redirect_stdout(output):
                app.main(input_folder)

            result = output.getvalue()
            self.assertIn("文件：broken.xlsx", result)
            self.assertIn(
                "读取失败：该文件可能已损坏，或不是有效的 Excel xlsx 文件。",
                result,
            )
            self.assertIn("文件：valid.xlsx", result)
            self.assertIn("Sheet数量：1", result)


if __name__ == "__main__":
    unittest.main()
